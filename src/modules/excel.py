import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict, Any


def generate_excel_report(data: List[Dict[str, Any]]) -> bytes:
    """
    Generates an XLSX report in memory from a list of dictionaries.

    Args:
        data (List[Dict[str, Any]]): Data to include in the report. Each dict represents a row.

    Returns:
        bytes: The binary content of the generated Excel file.
    """
    if not data:
        raise ValueError("Data list is empty.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Create headers automatically from the keys of the first dict
    headers = list(data[0].keys())
    ws.append(headers)

    # Insert data rows
    for row_dict in data:
        row_values = [row_dict.get(header, "") for header in headers]
        ws.append(row_values)

    # Format headers
    _apply_headers_format(ws)

    # Insert autofilter
    ws.auto_filter.ref = ws.dimensions

    # Automatically adjust column widths
    _auto_adjust_width(ws)

    # Save to memory and return bytes
    buffer = BytesIO()
    wb.save(buffer)

    return buffer.getvalue()

def _apply_headers_format(ws) -> None:
    """
    Applies formatting to the header row of the worksheet.
    
    Args:
        ws: The worksheet to format headers for.
    """
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

def _auto_adjust_width(ws) -> None:
    """
    Automatically adjusts column widths based on content.
    
    Args:
        ws: The worksheet to adjust column widths for.
    """
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                max_length = max(max_length, length)
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
